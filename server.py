from flask import Flask, request, jsonify, send_from_directory
import json, os, hashlib, csv, io
import openpyxl

app = Flask(__name__, static_folder='public')
DATA_FILE = os.path.join(os.path.dirname(__file__), 'data.json')

ADMIN_PASSWORD = hashlib.sha256(b'admin123').hexdigest()

DEFAULT_DATA = {
    "players": [],
    "matches": [],
    "predictions": {}
}

# Use PostgreSQL if DATABASE_URL is set, otherwise fall back to local file
DATABASE_URL = os.environ.get('DATABASE_URL')

def get_db():
    import psycopg2
    url = DATABASE_URL
    if url and url.startswith('postgres://'):
        url = url.replace('postgres://', 'postgresql://', 1)
    return psycopg2.connect(url)

def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute('''CREATE TABLE IF NOT EXISTS store (
        key TEXT PRIMARY KEY,
        value TEXT NOT NULL
    )''')
    conn.commit()
    cur.close()
    conn.close()

def load():
    if not DATABASE_URL:
        if not os.path.exists(DATA_FILE):
            save(DEFAULT_DATA)
        with open(DATA_FILE) as f:
            return json.load(f)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT value FROM store WHERE key = 'data'")
    row = cur.fetchone()
    cur.close()
    conn.close()
    if row:
        return json.loads(row[0])
    save(DEFAULT_DATA)
    return DEFAULT_DATA

def save(data):
    if not DATABASE_URL:
        with open(DATA_FILE, 'w') as f:
            json.dump(data, f, indent=2)
        return
    conn = get_db()
    cur = conn.cursor()
    cur.execute('''INSERT INTO store (key, value) VALUES ('data', %s)
                   ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value''',
                (json.dumps(data),))
    conn.commit()
    cur.close()
    conn.close()

if DATABASE_URL:
    with app.app_context():
        init_db()

def score_prediction(pred_home, pred_away, real_home, real_away):
    if real_home is None or real_away is None:
        return None
    if pred_home == real_home and pred_away == real_away:
        return 3
    pred_result = 'H' if pred_home > pred_away else ('A' if pred_home < pred_away else 'D')
    real_result = 'H' if real_home > real_away else ('A' if real_home < real_away else 'D')
    return 1 if pred_result == real_result else 0

@app.route('/')
def index():
    return send_from_directory('public', 'index.html')

@app.route('/api/data')
def get_data():
    data = load()
    predictions = data.get('predictions', {})
    matches = data.get('matches', [])
    players = data.get('players', [])

    # Build leaderboard
    starting_points = data.get('starting_points', {})
    leaderboard = {}
    for player in players:
        leaderboard[player] = {'name': player, 'points': starting_points.get(player, 0), 'exact': 0, 'correct': 0, 'played': 0}

    for match in matches:
        mid = match['id']
        rh, ra = match.get('real_home'), match.get('real_away')
        for player in players:
            preds = predictions.get(player, {})
            if mid in preds:
                ph, pa = preds[mid]['home'], preds[mid]['away']
                pts = score_prediction(ph, pa, rh, ra)
                if pts is not None:
                    leaderboard[player]['points'] += pts
                    leaderboard[player]['played'] += 1
                    if pts == 3:
                        leaderboard[player]['exact'] += 1
                    elif pts == 2:
                        leaderboard[player]['correct'] += 1

    leaderboard_list = sorted(leaderboard.values(), key=lambda x: (-x['points'], -x['exact'], x['name']))

    return jsonify({
        'players': players,
        'matches': matches,
        'predictions': predictions,
        'leaderboard': leaderboard_list
    })

@app.route('/api/admin/verify', methods=['POST'])
def verify_admin():
    body = request.json
    pw = hashlib.sha256(body.get('password', '').encode()).hexdigest()
    return jsonify({'ok': pw == ADMIN_PASSWORD})

@app.route('/api/admin/player', methods=['POST'])
def add_player():
    if not check_admin(request):
        return jsonify({'error': 'Unauthorized'}), 401
    data = load()
    name = request.json.get('name', '').strip()
    if not name or name in data['players']:
        return jsonify({'error': 'Invalid or duplicate name'}), 400
    data['players'].append(name)
    save(data)
    return jsonify({'ok': True})

@app.route('/api/admin/player/<name>', methods=['DELETE'])
def delete_player(name):
    if not check_admin(request):
        return jsonify({'error': 'Unauthorized'}), 401
    data = load()
    if name not in data['players']:
        return jsonify({'error': 'Not found'}), 404
    data['players'].remove(name)
    data['predictions'].pop(name, None)
    save(data)
    return jsonify({'ok': True})

@app.route('/api/admin/match', methods=['POST'])
def add_match():
    if not check_admin(request):
        return jsonify({'error': 'Unauthorized'}), 401
    data = load()
    body = request.json
    match_id = str(len(data['matches']) + 1) + '_' + body['home'] + '_' + body['away']
    match = {
        'id': match_id,
        'home': body['home'],
        'away': body['away'],
        'date': body.get('date', ''),
        'real_home': None,
        'real_away': None
    }
    data['matches'].append(match)
    save(data)
    return jsonify({'ok': True, 'match': match})

@app.route('/api/admin/match/<match_id>', methods=['DELETE'])
def delete_match(match_id):
    if not check_admin(request):
        return jsonify({'error': 'Unauthorized'}), 401
    data = load()
    data['matches'] = [m for m in data['matches'] if m['id'] != match_id]
    # Remove predictions for this match
    for player in data['predictions']:
        data['predictions'][player].pop(match_id, None)
    save(data)
    return jsonify({'ok': True})

@app.route('/api/admin/match/<match_id>/score', methods=['POST'])
def set_score(match_id):
    if not check_admin(request):
        return jsonify({'error': 'Unauthorized'}), 401
    data = load()
    body = request.json
    for match in data['matches']:
        if match['id'] == match_id:
            match['real_home'] = int(body['home'])
            match['real_away'] = int(body['away'])
            save(data)
            return jsonify({'ok': True})
    return jsonify({'error': 'Not found'}), 404

@app.route('/api/admin/score', methods=['POST'])
def set_score_by_id():
    if not check_admin(request):
        return jsonify({'error': 'Unauthorized'}), 401
    data = load()
    body = request.json
    match_id = body.get('match_id', '')
    for match in data['matches']:
        if match['id'] == match_id:
            match['real_home'] = int(body['home'])
            match['real_away'] = int(body['away'])
            save(data)
            return jsonify({'ok': True})
    return jsonify({'error': 'Not found'}), 404

@app.route('/api/admin/predict', methods=['POST'])
def admin_set_prediction():
    if not check_admin(request):
        return jsonify({'error': 'Unauthorized'}), 401
    data = load()
    body = request.json
    player = body.get('player', '').strip()
    match_id = body.get('match_id')
    ph = body.get('home')
    pa = body.get('away')
    if player not in data['players']:
        return jsonify({'error': 'Unknown player'}), 400
    if player not in data['predictions']:
        data['predictions'][player] = {}
    data['predictions'][player][match_id] = {'home': int(ph), 'away': int(pa)}
    save(data)
    return jsonify({'ok': True})

@app.route('/api/predict', methods=['POST'])
def predict():
    data = load()
    body = request.json
    player = body.get('player', '').strip()
    match_id = body.get('match_id')
    ph = body.get('home')
    pa = body.get('away')
    if player not in data['players']:
        return jsonify({'error': 'Unknown player'}), 400
    match = next((m for m in data['matches'] if m['id'] == match_id), None)
    if not match:
        return jsonify({'error': 'Unknown match'}), 400
    if match.get('real_home') is not None:
        return jsonify({'error': 'Match already played'}), 400
    if player not in data['predictions']:
        data['predictions'][player] = {}
    data['predictions'][player][match_id] = {'home': int(ph), 'away': int(pa)}
    save(data)
    return jsonify({'ok': True})

def parse_rows(file):
    """Parse CSV or XLSX file into list of row dicts (lowercased headers)."""
    filename = file.filename.lower()
    if filename.endswith('.xlsx'):
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip().lower() if h else '' for h in rows[0]]
        return [dict(zip(headers, [str(c).strip() if c is not None else '' for c in row])) for row in rows[1:]]
    else:
        text = file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        return [{k.strip().lower(): v.strip() for k, v in row.items()} for row in reader]

@app.route('/api/admin/import/players', methods=['POST'])
def import_players():
    if not check_admin(request):
        return jsonify({'error': 'Unauthorized'}), 401
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file'}), 400
    rows = parse_rows(f)
    data = load()
    added, skipped = [], []
    for row in rows:
        name = (row.get('name') or row.get('player') or next(iter(row.values()), '')).strip()
        if not name:
            continue
        if name in data['players']:
            skipped.append(name)
        else:
            data['players'].append(name)
            added.append(name)
    save(data)
    return jsonify({'ok': True, 'added': added, 'skipped': skipped})

@app.route('/api/admin/import/matches', methods=['POST'])
def import_matches():
    if not check_admin(request):
        return jsonify({'error': 'Unauthorized'}), 401
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file'}), 400
    rows = parse_rows(f)
    data = load()
    added, skipped = 0, 0
    for row in rows:
        home = (row.get('home') or row.get('home team') or row.get('hometeam') or '').strip()
        away = (row.get('away') or row.get('away team') or row.get('awayteam') or '').strip()
        date = (row.get('date') or '').strip()
        real_home_raw = (row.get('score home') or row.get('home score') or row.get('result home') or '').strip()
        real_away_raw = (row.get('score away') or row.get('away score') or row.get('result away') or '').strip()
        if not home or not away:
            skipped += 1
            continue
        existing = next((m for m in data['matches'] if m['home'] == home and m['away'] == away), None)
        if existing:
            # Update score if provided
            if real_home_raw.isdigit() and real_away_raw.isdigit():
                existing['real_home'] = int(real_home_raw)
                existing['real_away'] = int(real_away_raw)
            skipped += 1
            continue
        match_id = str(len(data['matches']) + 1) + '_' + home + '_' + away
        real_home = int(real_home_raw) if real_home_raw.isdigit() else None
        real_away = int(real_away_raw) if real_away_raw.isdigit() else None
        data['matches'].append({
            'id': match_id, 'home': home, 'away': away, 'date': date,
            'real_home': real_home, 'real_away': real_away
        })
        added += 1
    save(data)
    return jsonify({'ok': True, 'added': added, 'skipped': skipped})

PRED_ABBREVS = {
    'ivory': "côte d'ivoire", 'ger': 'germany', 'mex': 'mexico',
    'ecu': 'ecuador', 'arg': 'argentina', 'por': 'portugal',
    'uzb': 'uzbekistan', 'eng': 'england', 'cro': 'croatia',
    'nor': 'norway', 'sen': 'senegal', 'usa': 'united states',
    'bra': 'brazil', 'fra': 'france', 'esp': 'spain',
    'ned': 'netherlands', 'bel': 'belgium', 'uru': 'uruguay',
    'jpn': 'japan', 'jap': 'japan', 'aus': 'australia',
    'korea': 'korea republic', 'mor': 'morocco', 'mar': 'morocco',
    'sui': 'switzerland', 'swi': 'switzerland', 'pol': 'poland',
    'den': 'denmark', 'swe': 'sweden', 'ser': 'serbia',
    'nzl': 'new zealand', 'nz': 'new zealand', 'cam': 'cameroon',
    'tun': 'tunisia', 'sau': 'saudi arabia', 'ksa': 'saudi arabia',
    'col': 'colombia', 'jor': 'jordan', 'irq': 'iraq',
    'alg': 'algeria', 'aut': 'austria', 'pan': 'panama',
    'gha': 'ghana', 'crc': 'costa rica', 'pak': 'pakistan',
}

def _hints_team(word, team_name):
    import re as _re
    w = word.lower().strip('.,!?;:')
    t = team_name.lower()
    if len(w) < 2: return False
    if w in t or t.startswith(w): return True
    mapped = PRED_ABBREVS.get(w)
    if mapped and (t == mapped or t.startswith(mapped[:4])): return True
    return False

def _parse_pred_cell(cell, home, away):
    import re as _re, datetime as _dt
    if cell is None: return None
    if isinstance(cell, _dt.datetime):
        return (cell.month, cell.day)
    s = str(cell).strip().lstrip('`').strip()
    s = _re.sub(r'(?i)^draw\s*', '', s).strip()
    nums = _re.findall(r'\d+', s)
    if len(nums) < 2: return None
    n1, n2 = int(nums[0]), int(nums[1])
    words = _re.findall(r'[a-zA-Z\.]+', s)
    home_hit = any(_hints_team(w, home) for w in words)
    away_hit = any(_hints_team(w, away) for w in words)
    if away_hit and not home_hit:
        return (n2, n1)
    return (n1, n2)

@app.route('/api/admin/import/predictions', methods=['POST'])
def import_predictions():
    if not check_admin(request):
        return jsonify({'error': 'Unauthorized'}), 401
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file'}), 400
    if not f.filename.lower().endswith('.xlsx'):
        return jsonify({'error': 'Only .xlsx files supported for predictions'}), 400
    wb = openpyxl.load_workbook(f, data_only=True)
    data = load()
    by_teams = {(m['home'].lower(), m['away'].lower()): m for m in data['matches']}
    updated, skipped_players, skipped_preds = 0, [], 0
    for sheet_name in wb.sheetnames:
        player = sheet_name.strip()
        if player not in data['players']:
            skipped_players.append(player)
            continue
        ws = wb[sheet_name]
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row[1] or not row[2]: continue
            home_xl = str(row[1]).strip()
            away_xl = str(row[2]).strip()
            match = by_teams.get((home_xl.lower(), away_xl.lower()))
            if not match: continue
            pred = _parse_pred_cell(row[4], home_xl, away_xl)
            if pred is None:
                skipped_preds += 1
                continue
            nh, na = pred
            if player not in data['predictions']:
                data['predictions'][player] = {}
            data['predictions'][player][match['id']] = {'home': nh, 'away': na}
            updated += 1
    save(data)
    return jsonify({'ok': True, 'updated': updated,
                    'skipped_players': skipped_players, 'skipped_preds': skipped_preds})

def check_admin(req):
    pw = req.headers.get('X-Admin-Password', '')
    return hashlib.sha256(pw.encode()).hexdigest() == ADMIN_PASSWORD

if __name__ == '__main__':
    os.makedirs('public', exist_ok=True)
    print("Starting World Cup Predictor on http://localhost:8080")
    port = int(os.environ.get('PORT', 8080))
    app.run(debug=False, host='0.0.0.0', port=port)
